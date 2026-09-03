import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from database.connection import db
from config import DATA_DIR, COMPANY_NAME

from utils.retenciones_sri import calcular_retencion_sri

def export_sales_to_excel(date_from=None, date_to=None, output_path=None):
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(DATA_DIR, f"Reporte_Ventas_Inego_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    ws.views.sheetView[0].showGridLines = True

    navy_fill = PatternFill(start_color="0A192F", end_color="0A192F", fill_type="solid")
    blue_header_fill = PatternFill(start_color="1E3E7A", end_color="1E3E7A", fill_type="solid")
    zebra_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"REPORTE DE VENTAS - {COMPANY_NAME.upper()}"
    title_cell.font = title_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    sub_cell = ws["A2"]
    periodo_str = f"Desde: {date_from or 'Inicio'} | Hasta: {date_to or 'Actualidad'} | Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sub_cell.value = periodo_str
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="475569")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = [
        "N° Cotización", "Fecha Emisión", "Cliente / Razón Social", "RUC / Cédula", "Tipo Cliente",
        "Días Crédito", "Estado", "Aplica Retención SRI", "Producto / Ítem (Lista Desglosada)",
        "Cant.", "P. Unit ($)", "Subtotal Ítem ($)", "Subtotal Trans. ($)", "IVA 15% ($)",
        "Total Facturado ($)", "Ret. IR 1.75% ($)", "Ret. IVA 30% ($)", "Total Retenciones ($)", "Valor Neto a Cobrar ($)"
    ]
    ws.row_dimensions[4].height = 28

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = blue_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    query = """
        SELECT c.id AS cotizacion_id, c.numero_cotizacion, c.fecha_emision, cl.razon_social_nombre, 
               cl.ruc_cedula, cl.tipo_cliente, cl.dias_credito, c.estado, c.subtotal, c.iva, c.total
        FROM cotizaciones c
        JOIN clientes cl ON c.cliente_id = cl.id
    """
    params = []
    where_clauses = []
    if date_from:
        where_clauses.append("c.fecha_emision >= %s")
        params.append(date_from)
    if date_to:
        where_clauses.append("c.fecha_emision <= %s")
        params.append(date_to)
        
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
    query += " ORDER BY c.fecha_emision DESC"

    records = db.fetch_all(query, tuple(params))
    current_row = 5

    def _calc_ret(ruc_ced, sub, iv):
        res = calcular_retencion_sri(ruc_ced, sub, iv)
        return {
            'is_ruc': res['is_ruc'],
            'ap_txt': res['aplica_txt'],
            'r_ir': res['retencion_ir'],
            'r_iva': res['retencion_iva'],
            'tot_r': res['total_retenciones'],
            'tf': res['total_facturado'],
            'nc': res['neto_cobrar']
        }

    for idx, r in enumerate(records):
        sb = float(r["subtotal"] or 0.0)
        iv = float(r["iva"] or 0.0)
        ret = _calc_ret(r["ruc_cedula"], sb, iv)

        items_rows = db.fetch_all("""
            SELECT cd.cantidad, cd.precio_venta_unitario, cd.subtotal_linea, p.codigo, p.nombre
            FROM cotizacion_detalles cd
            JOIN productos p ON cd.producto_id = p.id
            WHERE cd.cotizacion_id = %s
        """, (r["cotizacion_id"],)) or []

        if not items_rows:
            items_rows = [{'cantidad': 1, 'precio_venta_unitario': sb, 'subtotal_linea': sb, 'codigo': 'PROD-GEN', 'nombre': 'Ítem Comercial General'}]

        for item in items_rows:
            ws.row_dimensions[current_row].height = 22
            ws.cell(row=current_row, column=1, value=r["numero_cotizacion"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=2, value=str(r["fecha_emision"])).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=3, value=r["razon_social_nombre"])
            ws.cell(row=current_row, column=4, value=r["ruc_cedula"] or "Consumidor Final").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=5, value=r["tipo_cliente"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=6, value=f"{r['dias_credito']} días" if r["tipo_cliente"] == "B2B" else "0 (Contado)").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=7, value=r["estado"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=8, value=ret["ap_txt"]).alignment = Alignment(horizontal="center")

            ws.cell(row=current_row, column=9, value=f"• {item['nombre']} ({item['codigo']})").alignment = Alignment(horizontal="left")
            ws.cell(row=current_row, column=10, value=item['cantidad']).alignment = Alignment(horizontal="center")
            
            c_pu = ws.cell(row=current_row, column=11, value=float(item['precio_venta_unitario']))
            c_pu.number_format = '$#,##0.00'
            
            c_si = ws.cell(row=current_row, column=12, value=float(item['subtotal_linea']))
            c_si.number_format = '$#,##0.00'

            c_sub = ws.cell(row=current_row, column=13, value=sb)
            c_sub.number_format = '$#,##0.00'

            c_iva = ws.cell(row=current_row, column=14, value=iv)
            c_iva.number_format = '$#,##0.00'

            c_tot = ws.cell(row=current_row, column=15, value=ret["tf"])
            c_tot.number_format = '$#,##0.00'

            c_rir = ws.cell(row=current_row, column=16, value=ret["r_ir"])
            c_rir.number_format = '$#,##0.00'

            c_riva = ws.cell(row=current_row, column=17, value=ret["r_iva"])
            c_riva.number_format = '$#,##0.00'

            c_tret = ws.cell(row=current_row, column=18, value=ret["tot_r"])
            c_tret.number_format = '$#,##0.00'
            if ret['is_ruc']:
                c_tret.font = Font(name="Calibri", size=11, bold=True, color="C2410C")

            c_net = ws.cell(row=current_row, column=19, value=ret["nc"])
            c_net.number_format = '$#,##0.00'
            c_net.font = bold_font

            fill_to_use = zebra_fill if idx % 2 == 1 else None
            for col_i in range(1, 20):
                c = ws.cell(row=current_row, column=col_i)
                c.border = thin_border
                if fill_to_use:
                    c.fill = fill_to_use

            current_row += 1

    ws.row_dimensions[current_row].height = 26
    ws.merge_cells(f"A{current_row}:L{current_row}")
    tot_lbl = ws.cell(row=current_row, column=1, value="TOTAL GENERAL:")
    tot_lbl.font = bold_font
    tot_lbl.alignment = Alignment(horizontal="right", vertical="center")

    for col_k in range(1, 13):
        ws.cell(row=current_row, column=col_k).border = thin_border
        ws.cell(row=current_row, column=col_k).fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")

    col_map_sales = {
        13: f"=SUM(M5:M{current_row-1})",  # Subtotal Trans.
        14: f"=SUM(N5:N{current_row-1})",  # IVA 15%
        15: f"=SUM(O5:O{current_row-1})",  # Total Facturado
        16: f"=SUM(P5:P{current_row-1})",  # Ret IR 1.75%
        17: f"=SUM(Q5:Q{current_row-1})",  # Ret IVA 30%
        18: f"=SUM(R5:R{current_row-1})",  # Total Retenciones
        19: f"=SUM(S5:S{current_row-1})"   # Neto a Cobrar
    }

    for col_idx, formula in col_map_sales.items():
        c_tot_final = ws.cell(row=current_row, column=col_idx, value=formula)
        c_tot_final.font = bold_font
        c_tot_final.number_format = '$#,##0.00'
        c_tot_final.border = thin_border
        c_tot_final.fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")

    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_s = str(cell.value or '')
            max_len = max(max_len, len(val_s))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 13)

    wb.save(output_path)
    return output_path


def export_gantt_chart_to_excel(output_path=None):
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(DATA_DIR, f"Diagrama_Gantt_Contratos_Gobierno_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Contratos Gobierno"
    ws.views.sheetView[0].showGridLines = True

    navy_fill = PatternFill(start_color="0A192F", end_color="0A192F", fill_type="solid")
    header_fill = PatternFill(start_color="1E3E7A", end_color="1E3E7A", fill_type="solid")
    bar_completed = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    bar_in_progress = PatternFill(start_color="38BDF8", end_color="38BDF8", fill_type="solid")
    bar_planned = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.merge_cells("A1:T1")
    title = ws["A1"]
    title.value = f"DIAGRAMA DE GANTT: GESTIÓN DE CONTRATOS Y LOGÍSTICA - {COMPANY_NAME.upper()}"
    title.font = title_font
    title.fill = navy_fill
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws["A2"] = "Leyenda de Avance:"
    ws["A2"].font = bold_font
    
    ws["C2"] = "Completado"
    ws["C2"].fill = bar_completed
    ws["C2"].font = Font(color="FFFFFF", bold=True)
    ws["C2"].alignment = Alignment(horizontal="center")

    ws["E2"] = "En Proceso"
    ws["E2"].fill = bar_in_progress
    ws["E2"].font = Font(color="000000", bold=True)
    ws["E2"].alignment = Alignment(horizontal="center")

    ws["G2"] = "Planificado"
    ws["G2"].fill = bar_planned
    ws["G2"].alignment = Alignment(horizontal="center")

    headers_base = ["ID Tarea", "Fase / Actividad Logística", "Responsable", "Inicio (Día)", "Duración (Días)", "% Avance"]
    for col_idx, h_text in enumerate(headers_base, 1):
        cell = ws.cell(row=4, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    start_date = datetime.now()
    for day_i in range(1, 15):
        col_idx = 6 + day_i
        curr_day = start_date + timedelta(days=day_i - 1)
        cell = ws.cell(row=4, column=col_idx, value=curr_day.strftime("%d/%m"))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 7

    tasks = [
        ("T-01", "Revisión de Pliegos y Requerimientos Contrato Gobierno", "Socio 1 - Dinero", 1, 2, 100),
        ("T-02", "Búsqueda y Búsqueda de Productos con Proveedores Guayaquil", "Socio 2 - Compras", 2, 3, 100),
        ("T-03", "Cotización de ítems importados (Amazon / Tiendamia)", "Socio 2 - Compras", 4, 2, 80),
        ("T-04", "Cuadro Comparativo de Precios y Selección de Proveedor", "Socio 1 & 2", 5, 2, 60),
        ("T-05", "Emisión de Orden de Compra y Pago de Anticipos", "Socio 3 - Contable", 7, 2, 40),
        ("T-06", "Recepción y Control de Calidad en Bodega Central", "Socio 2 - Compras", 9, 3, 20),
        ("T-07", "Consolidación de Paquetes y Logística de Entrega GYE", "Cualquier Socio", 11, 2, 0),
        ("T-08", "Firma de Acta Entrega-Recepción y Facturación 72 días", "Socio 3 - Contable", 13, 2, 0),
    ]

    row_start = 5
    for task_idx, t_data in enumerate(tasks):
        r = row_start + task_idx
        ws.row_dimensions[r].height = 24
        
        ws.cell(row=r, column=1, value=t_data[0]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=t_data[1])
        ws.cell(row=r, column=3, value=t_data[2])
        ws.cell(row=r, column=4, value=t_data[3]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5, value=t_data[4]).alignment = Alignment(horizontal="center")
        
        p_cell = ws.cell(row=r, column=6, value=f"{t_data[5]}%")
        p_cell.alignment = Alignment(horizontal="center")

        t_start = t_data[3]
        t_dur = t_data[4]
        t_prog = t_data[5]

        for col_i in range(1, 7):
            c = ws.cell(row=r, column=col_i)
            c.border = thin_border

        for day_i in range(1, 15):
            col_idx = 6 + day_i
            c = ws.cell(row=r, column=col_idx)
            c.border = thin_border
            
            if t_start <= day_i < (t_start + t_dur):
                if t_prog == 100:
                    c.fill = bar_completed
                elif t_prog > 0:
                    c.fill = bar_in_progress
                else:
                    c.fill = bar_planned

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12

    wb.save(output_path)
    return output_path


def export_expenses_to_excel(output_path=None):
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(DATA_DIR, f"Reporte_Gastos_CajaChica_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos Caja Chica"
    ws.views.sheetView[0].showGridLines = True

    navy_fill = PatternFill(start_color="0A192F", end_color="0A192F", fill_type="solid")
    blue_header_fill = PatternFill(start_color="1E3E7A", end_color="1E3E7A", fill_type="solid")
    zebra_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"REPORTE DE EGRESOS Y CAJA CHICA - {COMPANY_NAME.upper()}"
    title_cell.font = title_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = [
        "Fecha", "Categoría / Rubro", "Concepto / Descripción", 
        "Registrado Por", "Método de Pago", "Monto USD ($)"
    ]
    ws.row_dimensions[3].height = 28

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = blue_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    query = "SELECT fecha, categoria, concepto, registrado_por, metodo_pago, monto FROM gastos ORDER BY fecha DESC"
    records = db.fetch_all(query)
    current_row = 4
    total_acumulado = 0.0

    for idx, r in enumerate(records):
        ws.cell(row=current_row, column=1, value=str(r["fecha"])).alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=2, value=r["categoria"])
        ws.cell(row=current_row, column=3, value=r["concepto"])
        ws.cell(row=current_row, column=4, value=r["registrado_por"])
        ws.cell(row=current_row, column=5, value=r["metodo_pago"]).alignment = Alignment(horizontal="center")
        
        m_c = ws.cell(row=current_row, column=6, value=float(r["monto"]))
        m_c.number_format = '$#,##0.00'
        total_acumulado += float(r["monto"])

        fill_to_use = zebra_fill if idx % 2 == 1 else None
        for col_i in range(1, 7):
            c = ws.cell(row=current_row, column=col_i)
            c.font = regular_font
            c.border = thin_border
            if fill_to_use:
                c.fill = fill_to_use

        current_row += 1

    ws.row_dimensions[current_row].height = 25
    ws.cell(row=current_row, column=5, value="TOTAL GENERAL:").font = bold_font
    ws.cell(row=current_row, column=5).alignment = Alignment(horizontal="right", vertical="center")
    
    tot_final = ws.cell(row=current_row, column=6, value=total_acumulado)
    tot_final.font = bold_font
    tot_final.number_format = '$#,##0.00'
    tot_final.border = thin_border
    tot_final.fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(output_path)
    return output_path
