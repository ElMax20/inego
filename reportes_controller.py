"""
=============================================================================
CONTROLADOR DE REPORTES Y EXPORTACIÓN A EXCEL - INEGO INDUSTRIAS CRM
=============================================================================
Arquitectura desacoplada: Lógica de negocio y generación de hojas de cálculo
utilizando PySide6 (QObject, @Slot, Signal) y openpyxl.
Totalmente resiliente sin dependencia obligatoria de pandas.

Requerimientos Implementados:
- Reporte Diario de Ventas
- Reporte Mensual Consolidado
- Reporte Comercial por Rango de Fechas Personalizado
- Reporte de Salidas de Caja Chica (Agrupado por Categorías)
=============================================================================
"""

import os
import urllib.parse
from datetime import datetime
import calendar

try:
    import pandas as pd
except ImportError:
    pd = None

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from PySide6.QtCore import QObject, Slot, Signal
from database.connection import db
from config import COMPANY_NAME
from utils.retenciones_sri import calcular_retencion_sri


class ReportesController(QObject):
    """
    Controlador Python expuesto a QML para la generación y exportación
    profesional de reportes analíticos y comerciales en formato Excel (.xlsx).
    """

    # Señales emitidas hacia QML para retroalimentación visual del usuario
    exportSuccess = Signal(str, str)  # (mensaje_exito, ruta_archivo_generado)
    exportError = Signal(str)         # (mensaje_error)

    def __init__(self, parent=None):
        super().__init__(parent)

    # -------------------------------------------------------------------------
    # UTILIDADES DE RUTAS Y ESTILOS OPENPYXL
    # -------------------------------------------------------------------------
    @staticmethod
    def _sanitizar_ruta(ruta: str) -> str:
        """
        Limpia y normaliza la ruta recibida desde el FileDialog de QML,
        removiendo prefijos de protocolo (file:///) y asegurando extensión .xlsx.
        """
        if not ruta:
            return ""

        # Eliminar prefijos URL de QML
        if ruta.startswith("file:///"):
            ruta = ruta[8:]
        elif ruta.startswith("file://"):
            ruta = ruta[7:]

        # Decodificar caracteres especiales (%20, etc.)
        ruta = urllib.parse.unquote(ruta)

        # En Windows: corregir formato /C:/... a C:/...
        if len(ruta) > 2 and ruta[0] == '/' and ruta[2] == ':':
            ruta = ruta[1:]

        ruta = os.path.normpath(ruta)

        if not ruta.lower().endswith(".xlsx"):
            ruta += ".xlsx"

        # Asegurar que el directorio de destino exista
        dir_padre = os.path.dirname(ruta)
        if dir_padre and not os.path.exists(dir_padre):
            os.makedirs(dir_padre, exist_ok=True)

        return ruta

    @staticmethod
    def _obtener_estilos_base():
        """
        Define la paleta corporativa y tipografía profesional para openpyxl.
        """
        estilos = {
            "title_fill": PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid"),  # Navy oscuro
            "subtitle_fill": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"),
            "header_fill": PatternFill(start_color="1E3E7A", end_color="1E3E7A", fill_type="solid"),  # Azul corporativo
            "bronze_fill": PatternFill(start_color="8B5A2B", end_color="8B5A2B", fill_type="solid"),  # Acento bronce
            "kpi_fill": PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"),
            "zebra_fill": PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"),
            "total_fill": PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"),
            
            "title_font": Font(name="Calibri", size=15, bold=True, color="FFFFFF"),
            "subtitle_font": Font(name="Calibri", size=10, italic=True, color="475569"),
            "header_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "bold_font": Font(name="Calibri", size=11, bold=True, color="0F172A"),
            "regular_font": Font(name="Calibri", size=11, color="1E293B"),
            "kpi_num_font": Font(name="Calibri", size=13, bold=True, color="0F766E"),
            
            "thin_border": Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            ),
            "total_border": Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='475569'),
                bottom=Side(style='double', color='0F172A')
            )
        }
        return estilos

    @staticmethod
    def _ajustar_anchos_columna(ws, min_width=13, max_width=50):
        """
        Ajusta automáticamente el ancho de todas las columnas según el contenido máximo.
        """
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                if cell.coordinate in ws.merged_cells:
                    continue
                val_str = str(cell.value or '')
                if '\n' in val_str:
                    lines = val_str.split('\n')
                    max_len = max(max_len, max(len(l) for l in lines))
                else:
                    max_len = max(max_len, len(val_str))
            
            ancho_calculado = max(max_len + 3, min_width)
            ws.column_dimensions[col_letter].width = min(ancho_calculado, max_width)

    # -------------------------------------------------------------------------
    # UTILIDADES DE RETENCIONES DE LEY SRI (ECUADOR) Y DESGLOSE DE PRODUCTOS
    # -------------------------------------------------------------------------
    @staticmethod
    def _calcular_retenciones_sri(ruc_cedula: str, subtotal: float, iva: float, regimen: str = "Régimen General", tipo_transaccion: str = "Bienes", es_especial: bool = False):
        """
        Calcula las retenciones de impuestos según la normativa oficial del SRI del Ecuador
        invocando el motor especializado utils.retenciones_sri.
        """
        res = calcular_retencion_sri(
            ruc_cedula, subtotal, iva,
            regimen_tributario=regimen,
            tipo_transaccion=tipo_transaccion,
            es_contribuyente_especial=es_especial
        )
        return {
            'is_ruc': res['is_ruc'],
            'aplica_txt': res['aplica_txt'],
            'ret_ir': res['retencion_ir'],
            'ret_iva': res['retencion_iva'],
            'total_ret': res['total_retenciones'],
            'total_facturado': res['total_facturado'],
            'neto_cobrar': res['neto_cobrar']
        }

    @staticmethod
    def _obtener_detalles_items(cotizacion_id, subtotal_default=0.0):
        """
        Obtiene el desglose individual de productos de una cotización/orden para presentarlo
        como lista ítem por ítem en lugar de una sola línea consolidada.
        """
        if not cotizacion_id:
            return [{
                'codigo': 'PROD-GEN',
                'nombre': 'Ítem / Servicio Comercial General',
                'cantidad': 1,
                'precio_unitario': subtotal_default,
                'subtotal_linea': subtotal_default
            }]

        query = """
            SELECT 
                p.codigo AS producto_codigo,
                p.nombre AS producto_nombre,
                cd.cantidad,
                cd.precio_venta_unitario,
                cd.subtotal_linea
            FROM cotizacion_detalles cd
            JOIN productos p ON cd.producto_id = p.id
            WHERE cd.cotizacion_id = %s
            ORDER BY cd.id ASC
        """
        rows = db.fetch_all(query, (cotizacion_id,))
        if not rows:
            return [{
                'codigo': 'PROD-GEN',
                'nombre': 'Ítem / Servicio Comercial General',
                'cantidad': 1,
                'precio_unitario': subtotal_default,
                'subtotal_linea': subtotal_default
            }]

        res = []
        for r in rows:
            res.append({
                'codigo': r.get('producto_codigo') or 'N/A',
                'nombre': r.get('producto_nombre') or 'Producto sin nombre',
                'cantidad': int(r.get('cantidad') or 1),
                'precio_unitario': float(r.get('precio_venta_unitario') or 0.0),
                'subtotal_linea': float(r.get('subtotal_linea') or 0.0)
            })
        return res

    # -------------------------------------------------------------------------
    # REPORTE DIARIO DE VENTAS
    # -------------------------------------------------------------------------
    @Slot(str, str)
    def exportarReporteDiario(self, fecha_str: str, ruta_destino: str):
        """
        Genera el listado detallado de transacciones de ventas ejecutadas en
        una fecha específica con desglose tipo lista y retenciones SRI para RUC.
        """
        try:
            ruta_final = self._sanitizar_ruta(ruta_destino)
            if not ruta_final:
                self.exportError.emit("⚠️ Ruta de guardado no válida.")
                return

            fecha_limpia = (fecha_str or "").strip()
            try:
                datetime.strptime(fecha_limpia, "%Y-%m-%d")
            except ValueError:
                self.exportError.emit("⚠️ Formato de fecha inválido. Utilice YYYY-MM-DD.")
                return

            query_diario = """
                SELECT 
                    ov.numero_orden AS id_transaccion,
                    COALESCE(strftime('%H:%M:%S', ov.fecha_orden), '10:30:00') AS hora,
                    c.razon_social_nombre AS cliente,
                    c.ruc_cedula,
                    CASE 
                        WHEN c.tipo_cliente = 'B2B' THEN 'Crédito B2B (72 días)'
                        ELSE 'Contado / Efectivo'
                    END AS metodo_pago,
                    ov.cotizacion_id,
                    ov.subtotal,
                    ov.iva,
                    ov.total
                FROM ordenes_venta ov
                JOIN clientes c ON ov.cliente_id = c.id
                WHERE date(ov.fecha_orden) = date(%s)

                UNION ALL

                SELECT 
                    c.numero_cotizacion AS id_transaccion,
                    '10:00:00' AS hora,
                    cl.razon_social_nombre AS cliente,
                    cl.ruc_cedula,
                    CASE 
                        WHEN c.es_credito_72dias = 1 THEN 'Crédito B2B (72 días)'
                        ELSE 'Contado / Efectivo'
                    END AS metodo_pago,
                    c.id AS cotizacion_id,
                    c.subtotal,
                    c.iva,
                    c.total
                FROM cotizaciones c
                JOIN clientes cl ON c.cliente_id = cl.id
                WHERE date(c.fecha_emision) = date(%s)
                  AND c.id NOT IN (SELECT cotizacion_id FROM ordenes_venta WHERE cotizacion_id IS NOT NULL)
                ORDER BY hora ASC
            """
            filas = db.fetch_all(query_diario, (fecha_limpia, fecha_limpia)) or []

            registros = []
            for r in filas:
                registros.append({
                    'id_transaccion': r.get('id_transaccion') or 'N/A',
                    'hora': r.get('hora') or '--:--',
                    'cliente': r.get('cliente') or 'Consumidor Final',
                    'ruc_cedula': r.get('ruc_cedula') or '',
                    'metodo_pago': r.get('metodo_pago') or 'Efectivo',
                    'cotizacion_id': r.get('cotizacion_id'),
                    'subtotal': float(r.get('subtotal') or 0.0),
                    'iva': float(r.get('iva') or 0.0),
                    'total': float(r.get('total') or 0.0)
                })

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte Diario"
            ws.views.sheetView[0].showGridLines = True
            estilos = self._obtener_estilos_base()

            # Encabezado corporativo
            ws.merge_cells("A1:P1")
            t_cell = ws["A1"]
            t_cell.value = f"REPORTE DIARIO DE VENTAS Y RETENCIONES SRI — {COMPANY_NAME.upper()}"
            t_cell.font = estilos["title_font"]
            t_cell.fill = estilos["title_fill"]
            t_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 36

            # Subtítulo con metadata
            ws.merge_cells("A2:P2")
            s_cell = ws["A2"]
            s_cell.value = f"Fecha de Operación: {fecha_limpia}  |  Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            s_cell.font = estilos["subtitle_font"]
            s_cell.fill = estilos["subtitle_fill"]
            s_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 20

            # Bloque de KPIs
            tot_transacciones = len(registros)
            subtotal_dia = sum(r['subtotal'] for r in registros)
            total_dia = sum(r['total'] for r in registros)
            ret_totales_dia = sum(self._calcular_retenciones_sri(r['ruc_cedula'], r['subtotal'], r['iva'])['total_ret'] for r in registros)
            neto_total_dia = sum(self._calcular_retenciones_sri(r['ruc_cedula'], r['subtotal'], r['iva'])['neto_cobrar'] for r in registros)

            ws.cell(row=4, column=2, value="Transacciones:").font = estilos["bold_font"]
            ws.cell(row=4, column=3, value=tot_transacciones).font = estilos["kpi_num_font"]
            ws.cell(row=4, column=4, value="Total Facturado:").font = estilos["bold_font"]
            c_tot_kpi = ws.cell(row=4, column=5, value=total_dia)
            c_tot_kpi.font = estilos["kpi_num_font"]
            c_tot_kpi.number_format = "$#,##0.00"
            ws.cell(row=4, column=6, value="Retenciones SRI (RUC):").font = estilos["bold_font"]
            c_ret_kpi = ws.cell(row=4, column=7, value=ret_totales_dia)
            c_ret_kpi.font = estilos["kpi_num_font"]
            c_ret_kpi.number_format = "$#,##0.00"
            ws.cell(row=4, column=8, value="Neto a Cobrar Día:").font = estilos["bold_font"]
            c_net_kpi = ws.cell(row=4, column=9, value=neto_total_dia)
            c_net_kpi.font = estilos["kpi_num_font"]
            c_net_kpi.number_format = "$#,##0.00"
            ws.row_dimensions[4].height = 24

            # Cabecera de la tabla
            headers = [
                "ID Transacción", "Hora", "Cliente / Razón Social", "RUC / Cédula", "Aplica Retención SRI",
                "Producto / Ítem (Lista Desglosada)", "Cant.", "P. Unit ($)", "Subtotal Ítem ($)",
                "Subtotal Trans. ($)", "IVA 15% ($)", "Total Facturado ($)",
                "Ret. IR 1.75% ($)", "Ret. IVA 30% ($)", "Total Retenciones ($)", "Valor Neto a Cobrar ($)"
            ]
            header_row = 6
            ws.row_dimensions[header_row].height = 28

            for col_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=h_text)
                cell.font = estilos["header_font"]
                cell.fill = estilos["header_fill"]
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = estilos["thin_border"]

            current_r = header_row + 1

            if not registros:
                ws.merge_cells(f"A{current_r}:P{current_r}")
                empty_cell = ws.cell(row=current_r, column=1, value="No se registraron transacciones comerciales en la fecha especificada.")
                empty_cell.font = estilos["subtitle_font"]
                empty_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[current_r].height = 28
                current_r += 1
            else:
                for idx, row in enumerate(registros):
                    ret = self._calcular_retenciones_sri(row['ruc_cedula'], row['subtotal'], row['iva'])
                    items = self._obtener_detalles_items(row['cotizacion_id'], row['subtotal'])

                    for item_idx, item in enumerate(items):
                        ws.row_dimensions[current_r].height = 22
                        ws.cell(row=current_r, column=1, value=str(row['id_transaccion'])).alignment = Alignment(horizontal="center", vertical="center")
                        ws.cell(row=current_r, column=2, value=str(row['hora'])).alignment = Alignment(horizontal="center", vertical="center")
                        ws.cell(row=current_r, column=3, value=str(row['cliente'])).alignment = Alignment(horizontal="left", vertical="center")
                        ws.cell(row=current_r, column=4, value=str(row['ruc_cedula'] or 'Consumidor Final')).alignment = Alignment(horizontal="center", vertical="center")

                        c_ap = ws.cell(row=current_r, column=5, value=ret['aplica_txt'])
                        c_ap.alignment = Alignment(horizontal="center", vertical="center")
                        c_ap.font = Font(name="Calibri", size=10, bold=ret['is_ruc'], color="0F766E" if ret['is_ruc'] else "64748B")

                        c_prod = ws.cell(row=current_r, column=6, value=f"• {item['nombre']} ({item['codigo']})")
                        c_prod.alignment = Alignment(horizontal="left", vertical="center")
                        c_prod.font = estilos["regular_font"]

                        ws.cell(row=current_r, column=7, value=item['cantidad']).alignment = Alignment(horizontal="center", vertical="center")

                        c_pu = ws.cell(row=current_r, column=8, value=item['precio_unitario'])
                        c_pu.alignment = Alignment(horizontal="right", vertical="center")
                        c_pu.number_format = "$#,##0.00"

                        c_sub_i = ws.cell(row=current_r, column=9, value=item['subtotal_linea'])
                        c_sub_i.alignment = Alignment(horizontal="right", vertical="center")
                        c_sub_i.number_format = "$#,##0.00"

                        c_sub_t = ws.cell(row=current_r, column=10, value=row['subtotal'])
                        c_sub_t.alignment = Alignment(horizontal="right", vertical="center")
                        c_sub_t.number_format = "$#,##0.00"

                        c_iva = ws.cell(row=current_r, column=11, value=row['iva'])
                        c_iva.alignment = Alignment(horizontal="right", vertical="center")
                        c_iva.number_format = "$#,##0.00"

                        c_tf = ws.cell(row=current_r, column=12, value=ret['total_facturado'])
                        c_tf.alignment = Alignment(horizontal="right", vertical="center")
                        c_tf.number_format = "$#,##0.00"

                        c_rir = ws.cell(row=current_r, column=13, value=ret['ret_ir'])
                        c_rir.alignment = Alignment(horizontal="right", vertical="center")
                        c_rir.number_format = "$#,##0.00"

                        c_riva = ws.cell(row=current_r, column=14, value=ret['ret_iva'])
                        c_riva.alignment = Alignment(horizontal="right", vertical="center")
                        c_riva.number_format = "$#,##0.00"

                        c_tret = ws.cell(row=current_r, column=15, value=ret['total_ret'])
                        c_tret.alignment = Alignment(horizontal="right", vertical="center")
                        c_tret.number_format = "$#,##0.00"
                        if ret['is_ruc']:
                            c_tret.font = Font(name="Calibri", size=11, bold=True, color="C2410C")

                        c_net = ws.cell(row=current_r, column=16, value=ret['neto_cobrar'])
                        c_net.alignment = Alignment(horizontal="right", vertical="center")
                        c_net.number_format = "$#,##0.00"
                        c_net.font = estilos["bold_font"]

                        for col_k in range(1, 17):
                            cell_k = ws.cell(row=current_r, column=col_k)
                            cell_k.border = estilos["thin_border"]
                            if idx % 2 == 1:
                                cell_k.fill = estilos["zebra_fill"]

                        current_r += 1

                # Fila de Totales con fórmulas openpyxl
                ws.row_dimensions[current_r].height = 26
                ws.merge_cells(f"A{current_r}:H{current_r}")
                tot_lbl = ws.cell(row=current_r, column=1, value="TOTAL GENERAL DEL DÍA:")
                tot_lbl.font = estilos["bold_font"]
                tot_lbl.alignment = Alignment(horizontal="right", vertical="center")
                tot_lbl.border = estilos["total_border"]
                tot_lbl.fill = estilos["total_fill"]

                for col_m in range(2, 9):
                    ws.cell(row=current_r, column=col_m).border = estilos["total_border"]
                    ws.cell(row=current_r, column=col_m).fill = estilos["total_fill"]

                col_map = {
                    9: f"=SUM(I{header_row + 1}:I{current_r - 1})",   # Subtotal Ítem
                    10: f"=SUM(J{header_row + 1}:J{current_r - 1})",  # Subtotal Trans.
                    11: f"=SUM(K{header_row + 1}:K{current_r - 1})",  # IVA 15%
                    12: f"=SUM(L{header_row + 1}:L{current_r - 1})",  # Total Facturado
                    13: f"=SUM(M{header_row + 1}:M{current_r - 1})",  # Ret IR 1.75%
                    14: f"=SUM(N{header_row + 1}:N{current_r - 1})",  # Ret IVA 30%
                    15: f"=SUM(O{header_row + 1}:O{current_r - 1})",  # Total Retenciones
                    16: f"=SUM(P{header_row + 1}:P{current_r - 1})"   # Neto a Cobrar
                }

                for col_idx, formula in col_map.items():
                    c_tot_cell = ws.cell(row=current_r, column=col_idx, value=formula)
                    c_tot_cell.font = estilos["bold_font"]
                    c_tot_cell.alignment = Alignment(horizontal="right", vertical="center")
                    c_tot_cell.border = estilos["total_border"]
                    c_tot_cell.fill = estilos["total_fill"]
                    c_tot_cell.number_format = "$#,##0.00"

            self._ajustar_anchos_columna(ws)
            wb.save(ruta_final)
            wb.close()

            self.exportSuccess.emit(
                f"✅ Reporte Diario ({fecha_limpia}) exportado con éxito.",
                ruta_final
            )
        except Exception as e:
            self.exportError.emit(f"❌ Error al generar Reporte Diario: {str(e)}")

    # -------------------------------------------------------------------------
    # REPORTE MENSUAL CONSOLIDADO
    # -------------------------------------------------------------------------
    @Slot(int, int, str)
    @Slot(str, str, str)
    def exportarReporteMensual(self, mes, anio, ruta_destino: str):
        """
        Exporta a Excel el consolidado acumulado de ventas e ingresos del mes
        con desglose tipo lista por producto y retenciones SRI para RUC.
        """
        try:
            ruta_final = self._sanitizar_ruta(ruta_destino)
            if not ruta_final:
                self.exportError.emit("⚠️ Ruta de guardado no válida.")
                return

            mes_int = int(mes)
            anio_int = int(anio)

            if mes_int < 1 or mes_int > 12 or anio_int < 2000 or anio_int > 2100:
                self.exportError.emit("⚠️ Mes o año especificado fuera de rango válido.")
                return

            MESES_ESPANOL = {
                1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
                5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
                9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
            }
            DIAS_ESPANOL = {
                0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves",
                4: "Viernes", 5: "Sábado", 6: "Domingo"
            }

            _, ultimo_dia = calendar.monthrange(anio_int, mes_int)
            fecha_inicio = f"{anio_int:04d}-{mes_int:02d}-01"
            fecha_fin = f"{anio_int:04d}-{mes_int:02d}-{ultimo_dia:02d}"
            nombre_mes = MESES_ESPANOL.get(mes_int, calendar.month_name[mes_int].capitalize())

            query_mes_unificado = """
                SELECT 
                    date(ov.fecha_orden) AS fecha,
                    ov.numero_orden,
                    c.razon_social_nombre AS cliente,
                    c.ruc_cedula,
                    c.tipo_cliente,
                    ov.cotizacion_id,
                    ov.subtotal,
                    ov.iva,
                    ov.total
                FROM ordenes_venta ov
                JOIN clientes c ON ov.cliente_id = c.id
                WHERE date(ov.fecha_orden) >= date(%s) AND date(ov.fecha_orden) <= date(%s)

                UNION ALL

                SELECT 
                    date(c.fecha_emision) AS fecha,
                    c.numero_cotizacion AS numero_orden,
                    cl.razon_social_nombre AS cliente,
                    cl.ruc_cedula,
                    cl.tipo_cliente,
                    c.id AS cotizacion_id,
                    c.subtotal,
                    c.iva,
                    c.total
                FROM cotizaciones c
                JOIN clientes cl ON c.cliente_id = cl.id
                WHERE date(c.fecha_emision) >= date(%s) AND date(c.fecha_emision) <= date(%s)
                  AND c.id NOT IN (SELECT cotizacion_id FROM ordenes_venta WHERE cotizacion_id IS NOT NULL)
                ORDER BY fecha ASC
            """
            filas = db.fetch_all(query_mes_unificado, (fecha_inicio, fecha_fin, fecha_inicio, fecha_fin)) or []

            registros = []
            diario_map = {}
            for r in filas:
                f_str = str(r.get('fecha') or '')
                sb = float(r.get('subtotal') or 0.0)
                iv = float(r.get('iva') or 0.0)
                tot = float(r.get('total') or 0.0)
                ruc = str(r.get('ruc_cedula') or '')
                ret = self._calcular_retenciones_sri(ruc, sb, iv)

                rec = {
                    'fecha': f_str,
                    'numero_orden': r.get('numero_orden') or 'N/A',
                    'cliente': r.get('cliente') or 'Consumidor Final',
                    'ruc_cedula': ruc,
                    'tipo_cliente': r.get('tipo_cliente') or 'B2C',
                    'cotizacion_id': r.get('cotizacion_id'),
                    'subtotal': sb,
                    'iva': iv,
                    'total': tot,
                    'retenciones': ret
                }
                registros.append(rec)

                if f_str not in diario_map:
                    diario_map[f_str] = {'fecha': f_str, 'transacciones': 0, 'subtotal': 0.0, 'iva': 0.0, 'total': 0.0, 'retenciones': 0.0, 'neto': 0.0}
                diario_map[f_str]['transacciones'] += 1
                diario_map[f_str]['subtotal'] += sb
                diario_map[f_str]['iva'] += iv
                diario_map[f_str]['total'] += tot
                diario_map[f_str]['retenciones'] += ret['total_ret']
                diario_map[f_str]['neto'] += ret['neto_cobrar']

            list_diario = sorted(diario_map.values(), key=lambda x: x['fecha'])

            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Consolidado Mensual"
            ws1.views.sheetView[0].showGridLines = True
            estilos = self._obtener_estilos_base()

            # Título y encabezados
            ws1.merge_cells("A1:G1")
            t_cell = ws1["A1"]
            t_cell.value = f"CONSOLIDADO MENSUAL DE INGRESOS Y RETENCIONES — {COMPANY_NAME.upper()}"
            t_cell.font = estilos["title_font"]
            t_cell.fill = estilos["title_fill"]
            t_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws1.row_dimensions[1].height = 36

            ws1.merge_cells("A2:G2")
            s_cell = ws1["A2"]
            s_cell.value = f"Período: {nombre_mes} {anio_int} ({fecha_inicio} al {fecha_fin})  |  Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            s_cell.font = estilos["subtitle_font"]
            s_cell.fill = estilos["subtitle_fill"]
            s_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws1.row_dimensions[2].height = 20

            # Tarjetas KPIs
            tot_tx = len(registros)
            tot_sub = sum(r['subtotal'] for r in registros)
            tot_iva = sum(r['iva'] for r in registros)
            tot_facturado = sum(r['total'] for r in registros)
            tot_retenciones = sum(r['retenciones']['total_ret'] for r in registros)
            tot_neto = sum(r['retenciones']['neto_cobrar'] for r in registros)

            ws1.cell(row=4, column=1, value="MÉTRICAS CLAVE DEL MES:").font = estilos["bold_font"]
            kpi_labels = ["Total Transacciones:", "Total Facturado:", "Retenciones SRI (RUC):", "Valor Neto a Cobrar:"]
            kpi_values = [tot_tx, tot_facturado, tot_retenciones, tot_neto]

            for i, (lbl, val) in enumerate(zip(kpi_labels, kpi_values), start=1):
                ws1.cell(row=5, column=i+1 if i > 1 else 2, value=lbl).font = estilos["bold_font"]
                cell_val = ws1.cell(row=6, column=i+1 if i > 1 else 2, value=val)
                cell_val.font = estilos["kpi_num_font"]
                if i > 1:
                    cell_val.number_format = "$#,##0.00"

            # Tabla de Resumen por Día
            ws1.cell(row=8, column=1, value="DESGLOSE Y EVOLUCIÓN CRONOLÓGICA POR DÍA:").font = estilos["bold_font"]
            headers_d = ["Fecha", "Día", "N° Transacciones", "Subtotal ($)", "IVA 15% ($)", "Total Facturado ($)", "Retenciones SRI ($)", "Neto a Cobrar ($)"]
            ws1.row_dimensions[9].height = 24

            for c_idx, h in enumerate(headers_d, 1):
                cell = ws1.cell(row=9, column=c_idx, value=h)
                cell.font = estilos["header_font"]
                cell.fill = estilos["header_fill"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = estilos["thin_border"]

            curr_r = 10
            if not list_diario:
                ws1.merge_cells(f"A{curr_r}:H{curr_r}")
                ws1.cell(row=curr_r, column=1, value="No se registraron ventas en este mes.").font = estilos["subtitle_font"]
                curr_r += 1
            else:
                for idx, r in enumerate(list_diario):
                    try:
                        fecha_obj = datetime.strptime(str(r['fecha']), "%Y-%m-%d")
                        dia_semana = DIAS_ESPANOL.get(fecha_obj.weekday(), "N/A")
                    except Exception:
                        dia_semana = "N/A"

                    ws1.cell(row=curr_r, column=1, value=str(r['fecha'])).alignment = Alignment(horizontal="center")
                    ws1.cell(row=curr_r, column=2, value=dia_semana).alignment = Alignment(horizontal="center")
                    ws1.cell(row=curr_r, column=3, value=r['transacciones']).alignment = Alignment(horizontal="center")

                    c_sub = ws1.cell(row=curr_r, column=4, value=r['subtotal'])
                    c_sub.number_format = "$#,##0.00"
                    c_iva = ws1.cell(row=curr_r, column=5, value=r['iva'])
                    c_iva.number_format = "$#,##0.00"
                    c_tot = ws1.cell(row=curr_r, column=6, value=r['total'])
                    c_tot.number_format = "$#,##0.00"
                    c_ret = ws1.cell(row=curr_r, column=7, value=r['retenciones'])
                    c_ret.number_format = "$#,##0.00"
                    c_net = ws1.cell(row=curr_r, column=8, value=r['neto'])
                    c_net.number_format = "$#,##0.00"
                    c_net.font = estilos["bold_font"]

                    for c_k in range(1, 9):
                        ws1.cell(row=curr_r, column=c_k).border = estilos["thin_border"]
                        if idx % 2 == 1:
                            ws1.cell(row=curr_r, column=c_k).fill = estilos["zebra_fill"]

                    curr_r += 1

                ws1.merge_cells(f"A{curr_r}:B{curr_r}")
                ws1.cell(row=curr_r, column=1, value="TOTAL DEL MES:").alignment = Alignment(horizontal="right")
                ws1.cell(row=curr_r, column=1).font = estilos["bold_font"]
                ws1.cell(row=curr_r, column=3, value=f"=SUM(C10:C{curr_r-1})").font = estilos["bold_font"]

                for c_idx, col_let in enumerate(["D", "E", "F", "G", "H"], 4):
                    c_f = ws1.cell(row=curr_r, column=c_idx, value=f"=SUM({col_let}10:{col_let}{curr_r-1})")
                    c_f.number_format = "$#,##0.00"
                    c_f.font = estilos["bold_font"]

                for c_k in range(1, 9):
                    ws1.cell(row=curr_r, column=c_k).border = estilos["total_border"]
                    ws1.cell(row=curr_r, column=c_k).fill = estilos["total_fill"]

            self._ajustar_anchos_columna(ws1)

            # Hoja 2: Detalle completo tipo lista por ítem con retenciones SRI
            ws2 = wb.create_sheet(title="Transacciones Detalladas")
            ws2.views.sheetView[0].showGridLines = True

            headers_t = [
                "Fecha", "N° Transacción", "Cliente", "RUC / Cédula", "Tipo Cliente", "Aplica Retención SRI",
                "Producto / Ítem (Lista Desglosada)", "Cant.", "P. Unit ($)", "Subtotal Ítem ($)",
                "Subtotal Trans. ($)", "IVA 15% ($)", "Total Facturado ($)",
                "Ret. IR 1.75% ($)", "Ret. IVA 30% ($)", "Total Retenciones ($)", "Valor Neto a Cobrar ($)"
            ]
            for col_idx, h in enumerate(headers_t, 1):
                cell = ws2.cell(row=1, column=col_idx, value=h)
                cell.font = estilos["header_font"]
                cell.fill = estilos["header_fill"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = estilos["thin_border"]

            row_idx = 2
            for idx, r in enumerate(registros):
                ret = r['retenciones']
                items = self._obtener_detalles_items(r['cotizacion_id'], r['subtotal'])

                for item in items:
                    ws2.cell(row=row_idx, column=1, value=str(r['fecha'])).alignment = Alignment(horizontal="center")
                    ws2.cell(row=row_idx, column=2, value=str(r['numero_orden'])).alignment = Alignment(horizontal="center")
                    ws2.cell(row=row_idx, column=3, value=str(r['cliente']))
                    ws2.cell(row=row_idx, column=4, value=str(r['ruc_cedula'] or 'Consumidor Final')).alignment = Alignment(horizontal="center")
                    ws2.cell(row=row_idx, column=5, value=str(r['tipo_cliente'])).alignment = Alignment(horizontal="center")

                    c_ap = ws2.cell(row=row_idx, column=6, value=ret['aplica_txt'])
                    c_ap.alignment = Alignment(horizontal="center")
                    c_ap.font = Font(name="Calibri", size=10, bold=ret['is_ruc'], color="0F766E" if ret['is_ruc'] else "64748B")

                    c_prod = ws2.cell(row=row_idx, column=7, value=f"• {item['nombre']} ({item['codigo']})")
                    c_prod.alignment = Alignment(horizontal="left")

                    ws2.cell(row=row_idx, column=8, value=item['cantidad']).alignment = Alignment(horizontal="center")

                    c_pu = ws2.cell(row=row_idx, column=9, value=item['precio_unitario'])
                    c_pu.number_format = "$#,##0.00"

                    c_si = ws2.cell(row=row_idx, column=10, value=item['subtotal_linea'])
                    c_si.number_format = "$#,##0.00"

                    c_st = ws2.cell(row=row_idx, column=11, value=r['subtotal'])
                    c_st.number_format = "$#,##0.00"

                    c_iv = ws2.cell(row=row_idx, column=12, value=r['iva'])
                    c_iv.number_format = "$#,##0.00"

                    c_tf = ws2.cell(row=row_idx, column=13, value=ret['total_facturado'])
                    c_tf.number_format = "$#,##0.00"

                    c_rir = ws2.cell(row=row_idx, column=14, value=ret['ret_ir'])
                    c_rir.number_format = "$#,##0.00"

                    c_riva = ws2.cell(row=row_idx, column=15, value=ret['ret_iva'])
                    c_riva.number_format = "$#,##0.00"

                    c_tr = ws2.cell(row=row_idx, column=16, value=ret['total_ret'])
                    c_tr.number_format = "$#,##0.00"
                    if ret['is_ruc']:
                        c_tr.font = Font(name="Calibri", size=11, bold=True, color="C2410C")

                    c_net = ws2.cell(row=row_idx, column=17, value=ret['neto_cobrar'])
                    c_net.number_format = "$#,##0.00"
                    c_net.font = estilos["bold_font"]

                    for c_k in range(1, 18):
                        cell_k = ws2.cell(row=row_idx, column=c_k)
                        cell_k.border = estilos["thin_border"]
                        if idx % 2 == 1:
                            cell_k.fill = estilos["zebra_fill"]

                    row_idx += 1

            self._ajustar_anchos_columna(ws2)

            wb.save(ruta_final)
            wb.close()

            self.exportSuccess.emit(
                f"✅ Reporte Consolidado Mensual ({nombre_mes} {anio_int}) generado con éxito.",
                ruta_final
            )
        except Exception as e:
            self.exportError.emit(f"❌ Error al generar Reporte Mensual: {str(e)}")

    # -------------------------------------------------------------------------
    # REPORTE POR RANGO DE FECHAS PERSONALIZADO
    # -------------------------------------------------------------------------
    @Slot(str, str, str)
    def exportarReporteRango(self, fecha_inicio: str, fecha_fin: str, ruta_destino: str):
        """
        Genera reporte comercial de ventas filtrado por el intervalo [fecha_inicio, fecha_fin]
        con desglose tipo lista y retenciones SRI para RUC.
        """
        try:
            ruta_final = self._sanitizar_ruta(ruta_destino)
            if not ruta_final:
                self.exportError.emit("⚠️ Ruta de guardado no válida.")
                return

            f_ini_clean = (fecha_inicio or "").strip()
            f_fin_clean = (fecha_fin or "").strip()

            try:
                dt_ini = datetime.strptime(f_ini_clean, "%Y-%m-%d")
                dt_fin = datetime.strptime(f_fin_clean, "%Y-%m-%d")
            except ValueError:
                self.exportError.emit("⚠️ Formato de fechas inválido. Debe ingresar YYYY-MM-DD.")
                return

            if dt_ini > dt_fin:
                self.exportError.emit("⚠️ La fecha inicial no puede ser posterior a la fecha final.")
                return

            query_rango_unificado = """
                SELECT 
                    date(ov.fecha_orden) AS fecha,
                    ov.numero_orden,
                    c.razon_social_nombre AS cliente,
                    c.ruc_cedula,
                    c.tipo_cliente,
                    CASE 
                        WHEN c.tipo_cliente = 'B2B' THEN 'Crédito 72 días'
                        ELSE 'Contado'
                    END AS condicion_pago,
                    ov.estado,
                    ov.cotizacion_id,
                    ov.subtotal,
                    ov.iva,
                    ov.total
                FROM ordenes_venta ov
                JOIN clientes c ON ov.cliente_id = c.id
                WHERE date(ov.fecha_orden) >= date(%s) AND date(ov.fecha_orden) <= date(%s)

                UNION ALL

                SELECT 
                    date(c.fecha_emision) AS fecha,
                    c.numero_cotizacion AS numero_orden,
                    cl.razon_social_nombre AS cliente,
                    cl.ruc_cedula,
                    cl.tipo_cliente,
                    CASE 
                        WHEN c.es_credito_72dias = 1 THEN 'Crédito 72 días'
                        ELSE 'Contado'
                    END AS condicion_pago,
                    c.estado,
                    c.id AS cotizacion_id,
                    c.subtotal,
                    c.iva,
                    c.total
                FROM cotizaciones c
                JOIN clientes cl ON c.cliente_id = cl.id
                WHERE date(c.fecha_emision) >= date(%s) AND date(c.fecha_emision) <= date(%s)
                  AND c.id NOT IN (SELECT cotizacion_id FROM ordenes_venta WHERE cotizacion_id IS NOT NULL)
                ORDER BY fecha ASC
            """
            filas = db.fetch_all(query_rango_unificado, (f_ini_clean, f_fin_clean, f_ini_clean, f_fin_clean)) or []

            registros = []
            for r in filas:
                registros.append({
                    'fecha': str(r.get('fecha') or ''),
                    'numero_orden': str(r.get('numero_orden') or 'N/A'),
                    'cliente': str(r.get('cliente') or 'Consumidor Final'),
                    'ruc_cedula': str(r.get('ruc_cedula') or ''),
                    'tipo_cliente': str(r.get('tipo_cliente') or 'B2C'),
                    'condicion_pago': str(r.get('condicion_pago') or 'Contado'),
                    'estado': str(r.get('estado') or 'Completada'),
                    'cotizacion_id': r.get('cotizacion_id'),
                    'subtotal': float(r.get('subtotal') or 0.0),
                    'iva': float(r.get('iva') or 0.0),
                    'total': float(r.get('total') or 0.0)
                })

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte Comercial Rango"
            ws.views.sheetView[0].showGridLines = True
            estilos = self._obtener_estilos_base()

            # Encabezado
            ws.merge_cells("A1:R1")
            t_cell = ws["A1"]
            t_cell.value = f"REPORTE COMERCIAL Y RETENCIONES POR INTERVALO — {COMPANY_NAME.upper()}"
            t_cell.font = estilos["title_font"]
            t_cell.fill = estilos["title_fill"]
            t_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 36

            ws.merge_cells("A2:R2")
            s_cell = ws["A2"]
            s_cell.value = f"Intervalo: Desde {f_ini_clean} Hasta {f_fin_clean}  |  Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            s_cell.font = estilos["subtitle_font"]
            s_cell.fill = estilos["subtitle_fill"]
            s_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 20

            headers = [
                "Fecha", "N° Comprobante", "Cliente", "RUC / Cédula", "Tipo", "Condición Pago",
                "Estado", "Aplica Retención SRI", "Producto / Ítem (Lista Desglosada)",
                "Cant.", "P. Unit ($)", "Subtotal Ítem ($)", "Subtotal Trans. ($)", "IVA 15% ($)",
                "Total Facturado ($)", "Ret. IR 1.75% ($)", "Ret. IVA 30% ($)", "Total Retenciones ($)", "Valor Neto a Cobrar ($)"
            ]
            header_row = 4
            ws.row_dimensions[header_row].height = 28

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=h)
                cell.font = estilos["header_font"]
                cell.fill = estilos["header_fill"]
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = estilos["thin_border"]

            curr_r = 5
            if not registros:
                ws.merge_cells(f"A{curr_r}:R{curr_r}")
                ws.cell(row=curr_r, column=1, value="No se encontraron registros comerciales en el intervalo de fechas indicado.").font = estilos["subtitle_font"]
                curr_r += 1
            else:
                for idx, r in enumerate(registros):
                    ret = self._calcular_retenciones_sri(r['ruc_cedula'], r['subtotal'], r['iva'])
                    items = self._obtener_detalles_items(r['cotizacion_id'], r['subtotal'])

                    for item in items:
                        ws.cell(row=curr_r, column=1, value=r['fecha']).alignment = Alignment(horizontal="center")
                        ws.cell(row=curr_r, column=2, value=r['numero_orden']).alignment = Alignment(horizontal="center")
                        ws.cell(row=curr_r, column=3, value=r['cliente'])
                        ws.cell(row=curr_r, column=4, value=r['ruc_cedula'] or 'Consumidor Final').alignment = Alignment(horizontal="center")
                        ws.cell(row=curr_r, column=5, value=r['tipo_cliente']).alignment = Alignment(horizontal="center")
                        ws.cell(row=curr_r, column=6, value=r['condicion_pago']).alignment = Alignment(horizontal="center")
                        ws.cell(row=curr_r, column=7, value=r['estado']).alignment = Alignment(horizontal="center")

                        c_ap = ws.cell(row=curr_r, column=8, value=ret['aplica_txt'])
                        c_ap.alignment = Alignment(horizontal="center")
                        c_ap.font = Font(name="Calibri", size=10, bold=ret['is_ruc'], color="0F766E" if ret['is_ruc'] else "64748B")

                        c_prod = ws.cell(row=curr_r, column=9, value=f"• {item['nombre']} ({item['codigo']})")
                        c_prod.alignment = Alignment(horizontal="left")

                        ws.cell(row=curr_r, column=10, value=item['cantidad']).alignment = Alignment(horizontal="center")

                        c_pu = ws.cell(row=curr_r, column=11, value=item['precio_unitario'])
                        c_pu.number_format = "$#,##0.00"

                        c_si = ws.cell(row=curr_r, column=12, value=item['subtotal_linea'])
                        c_si.number_format = "$#,##0.00"

                        c_sub = ws.cell(row=curr_r, column=13, value=r['subtotal'])
                        c_sub.number_format = "$#,##0.00"

                        c_iva = ws.cell(row=curr_r, column=14, value=r['iva'])
                        c_iva.number_format = "$#,##0.00"

                        c_tot = ws.cell(row=curr_r, column=15, value=ret['total_facturado'])
                        c_tot.number_format = "$#,##0.00"

                        c_rir = ws.cell(row=curr_r, column=16, value=ret['ret_ir'])
                        c_rir.number_format = "$#,##0.00"

                        c_riva = ws.cell(row=curr_r, column=17, value=ret['ret_iva'])
                        c_riva.number_format = "$#,##0.00"

                        c_tret = ws.cell(row=curr_r, column=18, value=ret['total_ret'])
                        c_tret.number_format = "$#,##0.00"
                        if ret['is_ruc']:
                            c_tret.font = Font(name="Calibri", size=11, bold=True, color="C2410C")

                        c_net = ws.cell(row=curr_r, column=19, value=ret['neto_cobrar'])
                        c_net.number_format = "$#,##0.00"
                        c_net.font = estilos["bold_font"]

                        for c_k in range(1, 20):
                            cell_k = ws.cell(row=curr_r, column=c_k)
                            cell_k.border = estilos["thin_border"]
                            if idx % 2 == 1:
                                cell_k.fill = estilos["zebra_fill"]

                        curr_r += 1

                ws.merge_cells(f"A{curr_r}:K{curr_r}")
                ws.cell(row=curr_r, column=1, value="TOTAL DEL PERÍODO:").alignment = Alignment(horizontal="right")
                ws.cell(row=curr_r, column=1).font = estilos["bold_font"]

                for c_k in range(1, 12):
                    ws.cell(row=curr_r, column=c_k).border = estilos["total_border"]
                    ws.cell(row=curr_r, column=c_k).fill = estilos["total_fill"]

                col_map_rango = {
                    12: f"=SUM(L{header_row + 1}:L{curr_r - 1})",  # Subtotal Ítem
                    13: f"=SUM(M{header_row + 1}:M{curr_r - 1})",  # Subtotal Trans.
                    14: f"=SUM(N{header_row + 1}:N{curr_r - 1})",  # IVA 15%
                    15: f"=SUM(O{header_row + 1}:O{curr_r - 1})",  # Total Facturado
                    16: f"=SUM(P{header_row + 1}:P{curr_r - 1})",  # Ret IR 1.75%
                    17: f"=SUM(Q{header_row + 1}:Q{curr_r - 1})",  # Ret IVA 30%
                    18: f"=SUM(R{header_row + 1}:R{curr_r - 1})",  # Total Retenciones
                    19: f"=SUM(S{header_row + 1}:S{curr_r - 1})"   # Neto a Cobrar
                }

                for col_idx, formula in col_map_rango.items():
                    c_tot_cell = ws.cell(row=curr_r, column=col_idx, value=formula)
                    c_tot_cell.font = estilos["bold_font"]
                    c_tot_cell.alignment = Alignment(horizontal="right", vertical="center")
                    c_tot_cell.border = estilos["total_border"]
                    c_tot_cell.fill = estilos["total_fill"]
                    c_tot_cell.number_format = "$#,##0.00"

            self._ajustar_anchos_columna(ws)
            wb.save(ruta_final)
            wb.close()

            self.exportSuccess.emit(
                f"✅ Reporte Comercial por Rango ({f_ini_clean} al {f_fin_clean}) generado con éxito.",
                ruta_final
            )
        except Exception as e:
            self.exportError.emit(f"❌ Error al generar Reporte por Rango: {str(e)}")

    # -------------------------------------------------------------------------
    # REPORTE DE SALIDAS DE CAJA CHICA (AGRUPADO POR CATEGORÍA)
    # -------------------------------------------------------------------------
    @Slot(str, str, str)
    def exportarReporteCajaChica(self, fecha_inicio: str, fecha_fin: str, ruta_destino: str):
        """
        Exporta a Excel el desglose de gastos y salidas de caja chica agrupado
        y subtotalizado por categoría.
        """
        try:
            ruta_final = self._sanitizar_ruta(ruta_destino)
            if not ruta_final:
                self.exportError.emit("⚠️ Ruta de guardado no válida.")
                return

            f_ini_clean = (fecha_inicio or "").strip()
            f_fin_clean = (fecha_fin or "").strip()

            try:
                dt_ini = datetime.strptime(f_ini_clean, "%Y-%m-%d")
                dt_fin = datetime.strptime(f_fin_clean, "%Y-%m-%d")
            except ValueError:
                self.exportError.emit("⚠️ Formato de fechas inválido. Utilice YYYY-MM-DD.")
                return

            if dt_ini > dt_fin:
                self.exportError.emit("⚠️ La fecha inicial no puede ser posterior a la fecha final.")
                return

            query_gastos = """
                SELECT 
                    id AS gasto_id,
                    fecha,
                    COALESCE(categoria, 'Otros') AS categoria,
                    concepto,
                    monto,
                    metodo_pago,
                    registrado_por
                FROM gastos
                WHERE date(fecha) >= date(%s) AND date(fecha) <= date(%s)
                ORDER BY categoria ASC, fecha ASC
            """
            filas = db.fetch_all(query_gastos, (f_ini_clean, f_fin_clean)) or []

            registros = []
            cat_map = {}
            for r in filas:
                c_nom = str(r.get('categoria') or 'Otros')
                m_val = float(r.get('monto') or 0.0)
                
                rec = {
                    'gasto_id': r.get('gasto_id') or 0,
                    'fecha': str(r.get('fecha') or ''),
                    'categoria': c_nom,
                    'concepto': str(r.get('concepto') or ''),
                    'monto': m_val,
                    'metodo_pago': str(r.get('metodo_pago') or 'Efectivo'),
                    'registrado_por': str(r.get('registrado_por') or 'Sistema')
                }
                registros.append(rec)

                if c_nom not in cat_map:
                    cat_map[c_nom] = {'categoria': c_nom, 'cantidad': 0, 'total_monto': 0.0}
                cat_map[c_nom]['cantidad'] += 1
                cat_map[c_nom]['total_monto'] += m_val

            total_general = sum(r['monto'] for r in registros)

            list_cat = sorted(cat_map.values(), key=lambda x: x['total_monto'], reverse=True)
            for c in list_cat:
                c['porcentaje'] = (c['total_monto'] / total_general * 100.0) if total_general > 0 else 0.0

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Salidas Caja Chica"
            ws.views.sheetView[0].showGridLines = True
            estilos = self._obtener_estilos_base()

            # Título superior
            ws.merge_cells("A1:G1")
            t_cell = ws["A1"]
            t_cell.value = f"CONTROL Y SALIDAS DE CAJA CHICA — {COMPANY_NAME.upper()}"
            t_cell.font = estilos["title_font"]
            t_cell.fill = estilos["title_fill"]
            t_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 36

            ws.merge_cells("A2:G2")
            s_cell = ws["A2"]
            s_cell.value = f"Período: {f_ini_clean} al {f_fin_clean}  |  Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            s_cell.font = estilos["subtitle_font"]
            s_cell.fill = estilos["subtitle_fill"]
            s_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 20

            # SECCIÓN 1: RESUMEN AGRUPADO POR CATEGORÍA
            ws.cell(row=4, column=1, value="1. CONSOLIDADO AGRUPADO POR CATEGORÍA:").font = estilos["bold_font"]
            
            cat_headers = ["Categoría / Rubro", "N° Salidas", "Subtotal Egresos ($)", "% del Total"]
            ws.row_dimensions[5].height = 24
            for col_idx, h in enumerate(cat_headers, 1):
                cell = ws.cell(row=5, column=col_idx, value=h)
                cell.font = estilos["header_font"]
                cell.fill = estilos["bronze_fill"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = estilos["thin_border"]

            curr_r = 6
            if not list_cat:
                ws.merge_cells(f"A{curr_r}:D{curr_r}")
                ws.cell(row=curr_r, column=1, value="No se registraron salidas de caja chica en este período.").font = estilos["subtitle_font"]
                curr_r += 1
            else:
                for idx, r in enumerate(list_cat):
                    ws.cell(row=curr_r, column=1, value=str(r['categoria'])).alignment = Alignment(horizontal="left")
                    ws.cell(row=curr_r, column=2, value=int(r['cantidad'])).alignment = Alignment(horizontal="center")
                    
                    c_m = ws.cell(row=curr_r, column=3, value=float(r['total_monto']))
                    c_m.number_format = "$#,##0.00"
                    c_m.font = estilos["bold_font"]
                    c_m.alignment = Alignment(horizontal="right")
                    
                    c_p = ws.cell(row=curr_r, column=4, value=float(r['porcentaje']) / 100.0)
                    c_p.number_format = "0.0%"
                    c_p.alignment = Alignment(horizontal="right")

                    for c_k in range(1, 5):
                        ws.cell(row=curr_r, column=c_k).border = estilos["thin_border"]
                        if idx % 2 == 1:
                            ws.cell(row=curr_r, column=c_k).fill = estilos["zebra_fill"]

                    curr_r += 1

                # Fila total de categorías
                ws.cell(row=curr_r, column=1, value="TOTAL SALIDAS CAJA CHICA:").font = estilos["bold_font"]
                ws.cell(row=curr_r, column=2, value=f"=SUM(B6:B{curr_r-1})").font = estilos["bold_font"]
                ws.cell(row=curr_r, column=2).alignment = Alignment(horizontal="center")

                c_tot_cat = ws.cell(row=curr_r, column=3, value=f"=SUM(C6:C{curr_r-1})")
                c_tot_cat.number_format = "$#,##0.00"
                c_tot_cat.font = estilos["bold_font"]
                c_tot_cat.alignment = Alignment(horizontal="right")

                c_tot_pct = ws.cell(row=curr_r, column=4, value="100.0%")
                c_tot_pct.font = estilos["bold_font"]
                c_tot_pct.alignment = Alignment(horizontal="right")

                for c_k in range(1, 5):
                    ws.cell(row=curr_r, column=c_k).border = estilos["total_border"]
                    ws.cell(row=curr_r, column=c_k).fill = estilos["total_fill"]

                curr_r += 2

            # SECCIÓN 2: DESGLOSE DETALLADO DE CADA TRANSACCIÓN
            ws.cell(row=curr_r, column=1, value="2. DESGLOSE DETALLADO DE COMPROBANTES:").font = estilos["bold_font"]
            curr_r += 1

            det_headers = ["ID", "Fecha", "Categoría", "Concepto / Descripción", "Método Pago", "Registrado Por", "Monto Salida ($)"]
            ws.row_dimensions[curr_r].height = 24
            for col_idx, h in enumerate(det_headers, 1):
                cell = ws.cell(row=curr_r, column=col_idx, value=h)
                cell.font = estilos["header_font"]
                cell.fill = estilos["header_fill"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = estilos["thin_border"]

            header_det_r = curr_r
            curr_r += 1

            if not registros:
                ws.merge_cells(f"A{curr_r}:G{curr_r}")
                ws.cell(row=curr_r, column=1, value="Sin movimientos.").font = estilos["subtitle_font"]
            else:
                for idx, r in enumerate(registros):
                    ws.cell(row=curr_r, column=1, value=f"EGR-{int(r['gasto_id']):04d}").alignment = Alignment(horizontal="center")
                    ws.cell(row=curr_r, column=2, value=str(r['fecha'])).alignment = Alignment(horizontal="center")
                    ws.cell(row=curr_r, column=3, value=str(r['categoria']))
                    ws.cell(row=curr_r, column=4, value=str(r['concepto']))
                    ws.cell(row=curr_r, column=5, value=str(r['metodo_pago'])).alignment = Alignment(horizontal="center")
                    ws.cell(row=curr_r, column=6, value=str(r['registrado_por']))

                    c_monto = ws.cell(row=curr_r, column=7, value=float(r['monto']))
                    c_monto.number_format = "$#,##0.00"
                    c_monto.font = estilos["bold_font"]
                    c_monto.alignment = Alignment(horizontal="right")

                    for c_k in range(1, 8):
                        ws.cell(row=curr_r, column=c_k).border = estilos["thin_border"]
                        if idx % 2 == 1:
                            ws.cell(row=curr_r, column=c_k).fill = estilos["zebra_fill"]

                    curr_r += 1

                ws.merge_cells(f"A{curr_r}:F{curr_r}")
                ws.cell(row=curr_r, column=1, value="TOTAL GENERAL DE SALIDAS DETALLADAS:").alignment = Alignment(horizontal="right")
                ws.cell(row=curr_r, column=1).font = estilos["bold_font"]

                for c_k in range(1, 7):
                    ws.cell(row=curr_r, column=c_k).border = estilos["total_border"]
                    ws.cell(row=curr_r, column=c_k).fill = estilos["total_fill"]

                c_fin_tot = ws.cell(row=curr_r, column=7, value=f"=SUM(G{header_det_r+1}:G{curr_r-1})")
                c_fin_tot.number_format = "$#,##0.00"
                c_fin_tot.font = estilos["bold_font"]
                c_fin_tot.border = estilos["total_border"]
                c_fin_tot.fill = estilos["total_fill"]
                c_fin_tot.alignment = Alignment(horizontal="right")

            self._ajustar_anchos_columna(ws)
            wb.save(ruta_final)
            wb.close()

            self.exportSuccess.emit(
                f"✅ Reporte de Salidas de Caja Chica ({f_ini_clean} al {f_fin_clean}) generado con éxito.",
                ruta_final
            )
        except Exception as e:
            self.exportError.emit(f"❌ Error al generar Reporte de Caja Chica: {str(e)}")
